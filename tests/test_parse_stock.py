import openpyxl
import pytest
from pipeline.parse_stock import parse_description, load_bike_rows, load_paa_rows


def test_parse_description_three_parts_is_model_size_color():
    result = parse_description("OOSTEN DA 24, 14A, L")
    assert result == {
        "model_name": "OOSTEN DA 24",
        "size_code": "14A",
        "color_code": "L",
        "variant_extra": None,
    }


def test_parse_description_multiword_model():
    result = parse_description("STRATTOS S2 700C DA, S1, 1L")
    assert result == {
        "model_name": "STRATTOS S2 700C DA",
        "size_code": "S1",
        "color_code": "1L",
        "variant_extra": None,
    }


def test_parse_description_two_parts_is_model_color_no_size():
    result = parse_description("BOTTLE PODIUM MDL B 710ML, B")
    assert result == {
        "model_name": "BOTTLE PODIUM MDL B 710ML",
        "size_code": None,
        "color_code": "B",
        "variant_extra": None,
    }


def test_parse_description_one_part_is_model_only():
    result = parse_description("TOOL SPOKE CUTTER C-216")
    assert result == {
        "model_name": "TOOL SPOKE CUTTER C-216",
        "size_code": None,
        "color_code": None,
        "variant_extra": None,
    }


def test_parse_description_more_than_three_parts_keeps_remainder_raw():
    result = parse_description("TIRE ONE 700X30C P, TLE, F, B")
    assert result == {
        "model_name": "TIRE ONE 700X30C P",
        "size_code": None,
        "color_code": None,
        "variant_extra": "TLE, F, B",
    }


def test_parse_description_rejects_empty_string():
    with pytest.raises(ValueError):
        parse_description("   ")


def test_load_bike_rows_filters_and_parses(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Warehouse", "Article Code", "Description", "Brand",
               "Merchandise Category", "Quantity", "Ordered Quantity", "Price"])
    ws.append(["Outlet", 503200001, "STRATTOS S4 700C DA, S1, Z", "POLYGON",
               "BIKE-ROAD DROP BAR", 1, 2, 10600000])
    ws.append(["Outlet", 503200002, "STRATTOS S4 700C DA, M, Z", "POLYGON",
               "BIKE-ROAD DROP BAR", 1, None, 10600000])
    ws.append(["Outlet", 999999999, "SOME SHOE, 42, BLK", "NIKE",
               "FOOTWEAR", 5, None, 500000])
    xlsx_path = tmp_path / "sample.xlsx"
    wb.save(xlsx_path)

    rows = load_bike_rows(str(xlsx_path))

    assert len(rows) == 2
    assert rows[0] == {
        "article_code": 503200001,
        "brand": "POLYGON",
        "category": "BIKE-ROAD DROP BAR",
        "warehouse": "Outlet",
        "quantity": 1,
        "ordered_quantity": 2,
        "price": 10600000,
        "model_name": "STRATTOS S4 700C DA",
        "size_code": "S1",
        "color_code": "Z",
        "variant_extra": None,
    }
    assert rows[1]["ordered_quantity"] is None


def test_load_bike_rows_skips_empty_description(tmp_path, capsys):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Warehouse", "Article Code", "Description", "Brand",
               "Merchandise Category", "Quantity", "Ordered Quantity", "Price"])
    ws.append(["Outlet", 1, "", "POLYGON",
               "BIKE-MTB RIGID FRAME", 1, None, 1000000])
    xlsx_path = tmp_path / "sample.xlsx"
    wb.save(xlsx_path)

    rows = load_bike_rows(str(xlsx_path))

    assert rows == []
    assert "1" in capsys.readouterr().out


def test_load_paa_rows_excludes_bike_categories_and_parses_variants(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Warehouse", "Article Code", "Description", "Brand",
               "Merchandise Category", "Quantity", "Ordered Quantity", "Price"])
    ws.append(["Outlet", 503200001, "STRATTOS S4 700C DA, S1, Z", "POLYGON",
               "BIKE-ROAD DROP BAR", 1, None, 10600000])
    ws.append(["Outlet", 737392031, "SHOES SH-RC903E, 440, W", "SHIMANO",
               "FOOTWEAR", 2, None, 4998000])
    ws.append(["Outlet", 721852, "TOOL SPOKE CUTTER C-216", "HOZAN",
               "TOOLS", 1, None, 498000])
    xlsx_path = tmp_path / "sample.xlsx"
    wb.save(xlsx_path)

    rows = load_paa_rows(str(xlsx_path))

    assert len(rows) == 2
    assert [r["article_code"] for r in rows] == [737392031, 721852]
    assert rows[0]["size_code"] == "440"
    assert rows[0]["color_code"] == "W"
    assert rows[1]["model_name"] == "TOOL SPOKE CUTTER C-216"
    assert rows[1]["size_code"] is None
    assert rows[1]["color_code"] is None
